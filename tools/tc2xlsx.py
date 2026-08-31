#!/usr/bin/env python3
"""
tc2xlsx.py - doc bang Markdown 12 cot (generated.md + extended.md moi API, summary.md, bug-report.md)
-> xuat excel/23127183_HW06_TestCases.xlsx (5 sheet).

Nguon DUY NHAT: generator/specs/*.mjs (qua generated.md/extended.md da sinh). Script nay CHI doc
Markdown, khong tinh toan gi them.

Chay: python tools/tc2xlsx.py
"""
import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Thieu openpyxl. Cai bang: pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent

APIS = [
    ("api-01-login", "API-01-login"),
    ("api-02-apply-coupon", "API-02-coupon"),
    ("api-03-product-update", "API-03-produpd"),
]

HEADER = ["TC ID", "Kỹ thuật", "Tham số & phân vùng", "Request", "Auth", "Query / Body",
          "Expected status", "Expected body / schema", "Căn cứ", "Nguồn", "Audit", "Kết quả"]

CELL_WIDTHS = [14, 16, 40, 28, 16, 55, 14, 45, 40, 8, 12, 16]


def clean(text):
    """Bo dinh dang Markdown khoi 1 o: **bold**, `code`, <br>."""
    if text is None:
        return ""
    text = text.strip()
    text = re.sub(r"\*\*(.*?)\*\*", r"\1", text)
    text = re.sub(r"`([^`]*)`", r"\1", text)
    text = text.replace("<br>", "\n").replace("<br/>", "\n")
    return text


def parse_tables(md_path):
    """Tim MOI bang bat dau bang dong '| TC ID |' trong 1 file .md, tra ve list[list[str]]."""
    if not md_path.exists():
        print(f"  !! khong tim thay {md_path}")
        return []
    lines = md_path.read_text(encoding="utf-8").splitlines()
    rows = []
    in_table = False
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("| TC ID |"):
            in_table = True
            continue
        if in_table:
            if stripped.startswith("|---"):
                continue
            if not stripped.startswith("|"):
                in_table = False
                continue
            cells = [clean(c) for c in stripped.strip("|").split("|")]
            if len(cells) >= len(HEADER):
                rows.append(cells[: len(HEADER)])
    return rows


def style_sheet(ws, nrows):
    for i, w in enumerate(CELL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    for c in range(1, len(HEADER) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADER))}{nrows + 1}"
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    for r in range(2, nrows + 2):
        for c in range(1, len(HEADER) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)
        result_cell = ws.cell(row=r, column=12)
        v = (result_cell.value or "").lower()
        if "pass" in v:
            result_cell.fill = pass_fill
        elif "fail" in v:
            result_cell.fill = fail_fill


def write_case_sheet(wb, title, rows):
    ws = wb.create_sheet(title=title[:31])
    ws.append(HEADER)
    for row in rows:
        ws.append(row)
    style_sheet(ws, len(rows))
    return len(rows)


def write_text_sheet(wb, title, md_path, col_width=110):
    ws = wb.create_sheet(title=title[:31])
    ws.column_dimensions["A"].width = col_width
    if not md_path.exists():
        ws.append([f"(chua co {md_path.name})"])
        return
    for line in md_path.read_text(encoding="utf-8").splitlines():
        ws.append([clean(line)])
    ws.column_dimensions["A"].alignment = Alignment(wrap_text=True)


def main():
    wb = Workbook()
    wb.remove(wb.active)

    summary_path = ROOT / "test-cases" / "test-summary" / "summary.md"
    write_text_sheet(wb, "Summary", summary_path, col_width=100)
    print(f"Summary: da ghi tu {summary_path.name}")

    total = 0
    for slug, sheet_name in APIS:
        base = ROOT / "test-cases" / slug
        gen_rows = parse_tables(base / "generated.md")
        own_rows = parse_tables(base / "extended.md")
        rows = gen_rows + own_rows
        if not rows:
            print(f"!! LOI: khong tim thay bang nao cho {slug} - kiem tra generated.md/extended.md")
            sys.exit(1)
        n = write_case_sheet(wb, sheet_name, rows)
        total += n
        print(f"{sheet_name}: {n} dong ({len(gen_rows)} AI + {len(own_rows)} SV)")

    bug_path = ROOT / "bug-report" / "bug-report.md"
    write_text_sheet(wb, "Bugs", bug_path, col_width=130)
    print(f"Bugs: da ghi tu {bug_path.name}")

    out_dir = ROOT / "excel"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / "23127183_HW06_TestCases.xlsx"
    wb.save(out_path)
    print(f"\nTong: {total} test case -> {out_path}")


if __name__ == "__main__":
    main()
